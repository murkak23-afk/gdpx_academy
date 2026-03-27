from __future__ import annotations

from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO

from aiogram import F, Router
from aiogram.exceptions import TelegramBadRequest
from aiogram.fsm.context import FSMContext
from aiogram.types import BufferedInputFile, CallbackQuery, InlineKeyboardButton, InlineKeyboardMarkup, Message
from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from src.database.models.enums import SubmissionStatus
from src.keyboards import CALLBACK_INLINE_BACK, REPLY_BTN_BACK
from src.keyboards.admin_hints import HINT_STATS
from src.keyboards.callbacks import CB_ADMIN_STATS_EXCEL, CB_ADMIN_STATS_PAGE, CB_ADMIN_STATS_VIEW, CB_NOOP
from src.services import AdminService, AdminStatsService
from src.utils.text_format import edit_message_text_safe, non_empty_plain

router = Router(name="admin-stats-router")

STATS_PAGE_SIZE = 8

_PERIOD_LABEL = {"day": "день (UTC с 00:00)", "week": "7 дней", "month": "30 дней"}


def _period_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="День",
                    callback_data=f"{CB_ADMIN_STATS_VIEW}:day",
                ),
                InlineKeyboardButton(
                    text="7 дней",
                    callback_data=f"{CB_ADMIN_STATS_VIEW}:week",
                ),
                InlineKeyboardButton(
                    text="30 дней",
                    callback_data=f"{CB_ADMIN_STATS_VIEW}:month",
                ),
            ],
            [
                InlineKeyboardButton(text="📊 Excel (день)", callback_data=f"{CB_ADMIN_STATS_EXCEL}:day"),
                InlineKeyboardButton(text="Excel (7д)", callback_data=f"{CB_ADMIN_STATS_EXCEL}:week"),
            ],
            [
                InlineKeyboardButton(text="Excel (30д)", callback_data=f"{CB_ADMIN_STATS_EXCEL}:month"),
            ],
            [InlineKeyboardButton(text=REPLY_BTN_BACK, callback_data=CALLBACK_INLINE_BACK)],
        ],
    )


def _payout_nav_keyboard(period: str, page: int, total: int) -> list[list[InlineKeyboardButton]]:
    max_page = max((total + STATS_PAGE_SIZE - 1) // STATS_PAGE_SIZE - 1, 0)
    page = min(max(page, 0), max_page)
    return [
        [
            InlineKeyboardButton(text="⬅️", callback_data=f"{CB_ADMIN_STATS_PAGE}:{period}:{max(page - 1, 0)}"),
            InlineKeyboardButton(text=f"{page + 1}/{max_page + 1}", callback_data=CB_NOOP),
            InlineKeyboardButton(text="➡️", callback_data=f"{CB_ADMIN_STATS_PAGE}:{period}:{min(page + 1, max_page)}"),
        ],
    ]


async def _build_full_stats_message(
    session: AsyncSession, period: str, payout_page: int
) -> tuple[str, InlineKeyboardMarkup]:
    svc = AdminStatsService(session=session)
    start, end = svc.period_bounds(period)
    incoming = await svc.count_incoming_submissions(start, end)
    acc = await svc.count_by_status_reviewed(SubmissionStatus.ACCEPTED, start, end)
    blk = await svc.count_by_status_reviewed(SubmissionStatus.BLOCKED, start, end)
    nas = await svc.count_by_status_reviewed(SubmissionStatus.NOT_A_SCAN, start, end)
    by_cat = await svc.accepted_by_category(start, end)
    avg_accept = await svc.avg_accept_amount(start, end)
    top_sellers = await svc.top_sellers_by_accept_amount(start, end, limit=5)
    payout_rows, payout_total = await svc.payout_rows_paginated(
        start, end, page=payout_page, page_size=STATS_PAGE_SIZE
    )

    pl = _PERIOD_LABEL.get(period, period)
    lines = [
        HINT_STATS,
        "",
        f"📈 Сводка — {pl}",
        f"Интервал UTC: {start.strftime('%Y-%m-%d %H:%M')} — {end.strftime('%Y-%m-%d %H:%M')}",
        "",
        f"Поступило товаров (создано): {incoming}",
        f"Зачёт (accepted): {acc}",
        f"Блок: {blk}",
        f"Не скан: {nas}",
        "",
        "Зачёт по операторам (категориям):",
    ]
    accepted_amount_total = sum((amt for _, _, amt in by_cat), start=0)
    avg_daily_paid = (accepted_amount_total / Decimal("30")).quantize(Decimal("0.01")) if period == "month" else (
        (accepted_amount_total / Decimal("7")).quantize(Decimal("0.01")) if period == "week" else accepted_amount_total
    )
    lines.append(f"Сумма зачёта (USDT): {accepted_amount_total}")
    lines.append(f"Средний чек симки (USDT): {avg_accept}")
    lines.append(f"Payout velocity (USDT/день): {avg_daily_paid}")
    if not by_cat:
        lines.append("  — нет данных")
    else:
        for title, cnt, amt in by_cat:
            lines.append(f"  • {title}: {cnt} шт, {amt} USDT")
    lines.extend(["", "Топ продавцов по сумме зачёта:"])
    if not top_sellers:
        lines.append("  — нет данных")
    else:
        for label, amount, cnt in top_sellers:
            lines.append(f"  • {label}: {amount} USDT ({cnt} симок)")
    lines.extend(["", "Выплаты продавцам (по @nickname, только с выплатами в периоде):", ""])
    if not payout_rows:
        lines.append("  — нет записей")
    else:
        for r in payout_rows:
            lines.append(f"  • {r['label']}: {r['total_paid']} USDT ({r['payout_count']} выпл.)")

    text = "\n".join(lines)
    if len(text) > 3900:
        text = text[:3880] + "\n…"

    kb_rows = _period_keyboard().inline_keyboard[:-1]
    kb_rows.extend(_payout_nav_keyboard(period, payout_page, payout_total))
    kb_rows.append([InlineKeyboardButton(text=REPLY_BTN_BACK, callback_data=CALLBACK_INLINE_BACK)])
    return text, InlineKeyboardMarkup(inline_keyboard=kb_rows)


async def send_stats_hub(message: Message, session: AsyncSession, state: FSMContext | None = None) -> None:
    """Точка входа: раздел «Статистика» из reply-меню."""

    text, kb = await _build_full_stats_message(session, "day", 0)
    plain = non_empty_plain(text)
    if state is not None and message.chat is not None:
        data = await state.get_data()
        mid = data.get("admin_last_panel_message_id")
        if mid is not None:
            try:
                await message.bot.edit_message_text(
                    chat_id=message.chat.id,
                    message_id=int(mid),
                    text=plain,
                    reply_markup=kb,
                )
                return
            except TelegramBadRequest:
                pass
    sent = await message.answer(plain, reply_markup=kb)
    if state is not None and sent:
        from src.handlers.admin_menu import _admin_store_panel_message

        await _admin_store_panel_message(state, sent)


@router.callback_query(F.data.startswith(f"{CB_ADMIN_STATS_VIEW}:"))
async def on_stats_view_period(callback: CallbackQuery, session: AsyncSession) -> None:
    if callback.from_user is None or callback.data is None:
        return
    if not await AdminService(session=session).can_access_payout_finance(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    period = callback.data.split(":")[-1]
    await callback.answer()
    text, kb = await _build_full_stats_message(session, period, 0)
    if callback.message is not None:
        await edit_message_text_safe(callback.message, text, reply_markup=kb)


@router.callback_query(F.data.startswith(f"{CB_ADMIN_STATS_PAGE}:"))
async def on_stats_payout_page(callback: CallbackQuery, session: AsyncSession) -> None:
    if callback.from_user is None or callback.data is None:
        return
    if not await AdminService(session=session).can_access_payout_finance(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    parts = callback.data.split(":")
    if len(parts) != 5:
        return
    period, page_s = parts[3], parts[4]
    page = int(page_s)
    await callback.answer()
    text, kb = await _build_full_stats_message(session, period, page)
    if callback.message is not None:
        await edit_message_text_safe(callback.message, text, reply_markup=kb)


@router.callback_query(F.data.startswith(f"{CB_ADMIN_STATS_EXCEL}:"))
async def on_stats_excel(callback: CallbackQuery, session: AsyncSession) -> None:
    if callback.from_user is None or callback.data is None:
        return
    if not await AdminService(session=session).can_access_payout_finance(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    period = callback.data.split(":")[-1]
    await callback.answer("Формирую файл…")

    svc = AdminStatsService(session=session)
    start, end = svc.period_bounds(period)
    pl = _PERIOD_LABEL.get(period, period)

    incoming = await svc.count_incoming_submissions(start, end)
    acc = await svc.count_by_status_reviewed(SubmissionStatus.ACCEPTED, start, end)
    blk = await svc.count_by_status_reviewed(SubmissionStatus.BLOCKED, start, end)
    nas = await svc.count_by_status_reviewed(SubmissionStatus.NOT_A_SCAN, start, end)
    by_cat = await svc.accepted_by_category(start, end)

    payout_rows, _ = await svc.payout_rows_paginated(start, end, page=0, page_size=5000)

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Сводка"
    ws0.append(["Период", pl])
    ws0.append(["UTC с", start.isoformat()])
    ws0.append(["UTC по", end.isoformat()])
    ws0.append([])
    ws0.append(["Поступило товаров", incoming])
    ws0.append(["Зачёт", acc])
    ws0.append(["Блок", blk])
    ws0.append(["Не скан", nas])

    ws1 = wb.create_sheet("По_операторам")
    ws1.append(["Категория (оператор)", "Зачёт, шт.", "Сумма зачёта, USDT"])
    for title, cnt, amt in by_cat:
        ws1.append([title, cnt, float(amt)])

    ws2 = wb.create_sheet("Выплаты")
    ws2.append(["Продавец", "Сумма USDT", "Число выплат"])
    for r in payout_rows:
        ws2.append([r["label"], float(r["total_paid"]), r["payout_count"]])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"stats_{period}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.xlsx"

    if callback.message is not None:
        await callback.message.answer_document(
            BufferedInputFile(buf.getvalue(), filename=fname),
            caption=f"Отчёт: {pl}",
        )
