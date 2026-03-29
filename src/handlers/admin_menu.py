from __future__ import annotations

import asyncio
import csv
import logging
import re
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from html import escape
from io import BytesIO, StringIO

from aiogram import Bot, F, Router
from aiogram.dispatcher.event.bases import SkipHandler
from aiogram.exceptions import TelegramAPIError, TelegramBadRequest
from aiogram.filters import Command, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.types import (
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Message,
    ReplyKeyboardRemove,
)
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload

from src.database.models.enums import PayoutStatus, SubmissionStatus
from src.database.models.publication import Payout
from src.database.models.submission import ReviewAction, Submission
from src.database.models.user import User
from src.handlers.moderation import on_moderation_queue
from src.keyboards import (
    BUTTON_ENTER_ADMIN_PANEL,
    BUTTON_EXIT_ADMIN_PANEL,
    CALLBACK_INLINE_BACK,
    REPLY_BTN_BACK,
    is_admin_main_menu_text,
    match_admin_menu_canonical,
    moderation_review_keyboard,
    pagination_keyboard,
    payout_confirm_keyboard,
    payout_final_confirm_keyboard,
    search_report_keyboard,
    seller_main_menu_keyboard,
)
from src.keyboards.admin_hints import (
    HINT_BROADCAST,
    HINT_PAYOUTS,
)
from src.keyboards.callbacks import (
    CB_ADMIN_BROADCAST,
    CB_ADMIN_INWORK_HUB,
    CB_ADMIN_INWORK_OPEN,
    CB_ADMIN_INWORK_PAGE,
    CB_ADMIN_INWORK_SEARCH,
    CB_ADMIN_PAYOUTS,
    CB_ADMIN_QUEUE,
    CB_ADMIN_QUEUE_START,
    CB_ADMIN_REPORT_SUBMISSION,
    CB_ADMIN_RESTRICT,
    CB_ADMIN_SEARCH_PAGE,
    CB_ADMIN_STATS_EXPORT_MONTH,
    CB_ADMIN_STATS_MONTH,
    CB_ADMIN_STATS_RESET,
    CB_ADMIN_STATS_RESET_CONFIRM,
    CB_ADMIN_DASHBOARD_RESET,
    CB_ADMIN_DASHBOARD_RESET_CONFIRM,
    CB_ADMIN_UNRESTRICT,
    CB_NOOP,
    CB_PAY_CANCEL,
    CB_PAY_CONFIRM,
    CB_PAY_FINAL_CONFIRM,
    CB_PAY_HISTORY_PAGE,
    CB_PAY_LEDGER_PAGE,
    CB_PAY_MARK,
    CB_PAY_PENDING_DELETE,
    CB_PAY_PENDING_PAGE,
    CB_PAY_TOPUP,
    CB_PAY_TOPUP_CHECK,
    CB_PAY_TRASH,
    CB_PAY_TRASH_PAGE,
)
from src.services import (
    AdminAuditService,
    AdminService,
    AdminStatsService,
    BillingService,
    CategoryService,
    CryptoBotService,
    SubmissionService,
    UserService,
)
from src.states.admin_state import AdminBroadcastState, AdminPayoutState, AdminSearchSimState
from src.states.moderation_state import AdminBatchPickState, AdminInReviewLookupState, AdminModerationForwardState
from src.utils.admin_keyboard import build_admin_main_inline_keyboard
from src.utils.admin_panel_text import ADMIN_PANEL_HOME_TEXT
from src.utils.submission_format import submission_status_emoji_line
from src.utils.submission_media import message_answer_submission
from src.utils.text_format import (
    edit_message_text_safe,
    non_empty_plain,
)
from src.utils.ui_builder import GDPXRenderer

router = Router(name="admin-router")
logger = logging.getLogger(__name__)
PHONE_QUERY_PATTERN = re.compile(r"^\+7\d{10}$")
PAGE_SIZE = 5
LEDGER_PAGE_SIZE = 8
INWORK_PAGE_SIZE = 8
_ADMIN_LAST_PANEL_MSG_KEY = "admin_last_panel_message_id"


async def _delete_message_later(bot: Bot, chat_id: int, message_id: int, delay_sec: int = 20) -> None:
    """Удаляет служебное сообщение с задержкой, чтобы не захламлять чат."""

    await asyncio.sleep(delay_sec)
    try:
        await bot.delete_message(chat_id=chat_id, message_id=message_id)
    except TelegramBadRequest:
        pass


def _month_stats_keyboard(year: int, month: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="📥 Выгрузить Excel",
                    callback_data=f"{CB_ADMIN_STATS_EXPORT_MONTH}:{year}:{month}",
                )
            ],
            [
                InlineKeyboardButton(
                    text="🗑 Обнулить статистику",
                    callback_data=CB_ADMIN_STATS_RESET,
                )
            ],
            [InlineKeyboardButton(text=REPLY_BTN_BACK, callback_data=CALLBACK_INLINE_BACK)],
        ]
    )


def _month_stats_text(rows: list[dict[str, int | object]], year: int, month: int) -> str:
    lines = [
        f"📊 <b>Статистика SIM за {month:02d}.{year} (UTC)</b>",
        "",
        "<code>День | Вход | ✅ | ❌ | 🚫 | ⛔</code>",
    ]

    total_incoming = 0
    total_accepted = 0
    total_rejected = 0
    total_blocked = 0
    total_not_scan = 0

    for row in rows:
        dt = row["date"]
        day = getattr(dt, "day", 0)
        incoming = int(row["incoming"])
        accepted = int(row["accepted"])
        rejected = int(row["rejected"])
        blocked = int(row["blocked"])
        not_scan = int(row["not_a_scan"])

        total_incoming += incoming
        total_accepted += accepted
        total_rejected += rejected
        total_blocked += blocked
        total_not_scan += not_scan

        lines.append(
            f"<code>{day:02d}   | {incoming:4d} | {accepted:3d} | {rejected:3d} | {blocked:3d} | {not_scan:3d}</code>"
        )

    total_failed = total_rejected + total_blocked + total_not_scan
    lines.extend(
        [
            "",
            "<b>Итого за месяц</b>",
            f"📥 Входящих SIM: <b>{total_incoming}</b>",
            f"✅ Принято: <b>{total_accepted}</b>",
            f"❌ Брак всего: <b>{total_failed}</b> (rejected={total_rejected}, blocked={total_blocked}, not_a_scan={total_not_scan})",
        ]
    )
    return "\n".join(lines)


def _month_stats_workbook_bytes(rows: list[dict[str, int | object]], year: int, month: int) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "SIM Daily"
    ws.append(["Дата", "Входящие", "Принято", "Rejected", "Blocked", "Not a scan", "Брак всего"])

    total_incoming = 0
    total_accepted = 0
    total_rejected = 0
    total_blocked = 0
    total_not_scan = 0

    for row in rows:
        dt = row["date"]
        incoming = int(row["incoming"])
        accepted = int(row["accepted"])
        rejected = int(row["rejected"])
        blocked = int(row["blocked"])
        not_scan = int(row["not_a_scan"])
        failed_total = rejected + blocked + not_scan

        total_incoming += incoming
        total_accepted += accepted
        total_rejected += rejected
        total_blocked += blocked
        total_not_scan += not_scan

        ws.append([
            str(dt),
            incoming,
            accepted,
            rejected,
            blocked,
            not_scan,
            failed_total,
        ])

    ws.append([])
    ws.append(["ИТОГО", total_incoming, total_accepted, total_rejected, total_blocked, total_not_scan, total_rejected + total_blocked + total_not_scan])

    meta = wb.create_sheet("Summary")
    meta.append(["Период", f"{month:02d}.{year} UTC"])
    meta.append(["Входящие SIM", total_incoming])
    meta.append(["Принято", total_accepted])
    meta.append(["Rejected", total_rejected])
    meta.append(["Blocked", total_blocked])
    meta.append(["Not a scan", total_not_scan])
    meta.append(["Брак всего", total_rejected + total_blocked + total_not_scan])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def _admin_delete_prev_panel_message(message: Message, state: FSMContext) -> None:
    data = await state.get_data()
    old_id = data.get(_ADMIN_LAST_PANEL_MSG_KEY)
    if old_id is None or message.chat is None:
        return
    try:
        await message.bot.delete_message(chat_id=message.chat.id, message_id=int(old_id))
    except TelegramBadRequest:
        pass


async def _admin_store_panel_message(state: FSMContext, sent: Message | None) -> None:
    if sent is None:
        return
    await state.update_data(**{_ADMIN_LAST_PANEL_MSG_KEY: sent.message_id})


def _lock_line(submission: Submission) -> str:
    if submission.locked_by_admin is None:
        return ""
    if submission.locked_by_admin.username:
        return f"🔒 ЗАБЛОКИРОВАНО: @{escape(submission.locked_by_admin.username)}"
    return f"🔒 ЗАБЛОКИРОВАНО: id:{submission.locked_by_admin.id}"


def _short_phone(value: str | None) -> str:
    text = (value or "").strip() or "—"
    return text if len(text) <= 24 else f"{text[:21]}..."


def _in_work_hub_keyboard(*, items: list[Submission], page: int, total: int) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    for item in items:
        rows.append(
            [
                InlineKeyboardButton(
                    text=f"📱 {_short_phone(item.description_text)}",
                    callback_data=f"{CB_ADMIN_INWORK_OPEN}:{item.id}",
                )
            ]
        )
    max_page = max((total - 1) // INWORK_PAGE_SIZE, 0) if total > 0 else 0
    page = min(max(page, 0), max_page)
    if total > INWORK_PAGE_SIZE:
        nav: list[InlineKeyboardButton] = []
        if page > 0:
            nav.append(InlineKeyboardButton(text="<<", callback_data=f"{CB_ADMIN_INWORK_PAGE}:{page - 1}"))
        nav.append(InlineKeyboardButton(text=f"{page + 1}/{max_page + 1}", callback_data=CB_NOOP))
        if page < max_page:
            nav.append(InlineKeyboardButton(text=">>", callback_data=f"{CB_ADMIN_INWORK_PAGE}:{page + 1}"))
        rows.append(nav)
    rows.append([InlineKeyboardButton(text="🔍 Найти по номеру", callback_data=CB_ADMIN_INWORK_SEARCH)])
    rows.append([InlineKeyboardButton(text=REPLY_BTN_BACK, callback_data=CALLBACK_INLINE_BACK)])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _parse_pay_uid_page(callback_data: str) -> tuple[int, int]:
    parts = callback_data.split(":")
    if len(parts) < 3:
        return 0, 0
    uid = int(parts[2])
    page = int(parts[3]) if len(parts) > 3 else 0
    return uid, page


def _pay_op_label(username: str) -> str:
    u = (username or "").strip() or "—"
    if len(u) > 40:
        u = u[:37] + "..."
    return f"Оплатить {u}"


async def _payout_ledger_text_and_markup(
    session: AsyncSession,
    *,
    page: int,
) -> tuple[str, InlineKeyboardMarkup]:
    # Получаем PENDING payouts из таблицы payouts
    rows, total = await BillingService(session=session).get_payouts_paginated(
        status=PayoutStatus.PENDING,
        page=page,
        page_size=LEDGER_PAGE_SIZE,
    )
    max_page = max((total - 1) // LEDGER_PAGE_SIZE, 0) if total > 0 else 0
    page = min(max(page, 0), max_page)

    lines = ["💰 ВЕДОМОСТЬ ВЫПЛАТ", ""]
    if not rows:
        lines.append("Нет пользователей с ожидающими выплатами.")
    else:
        for i, (payout, user) in enumerate(rows, start=page * LEDGER_PAGE_SIZE + 1):
            username = f"@{user.username}" if user.username else f"@{user.telegram_id}"
            lines.append(f"{i}. {username} | {payout.accepted_count} шт. | {payout.amount} USDT")
    text = "\n".join(lines)
    kb_rows: list[list[InlineKeyboardButton]] = []
    for payout, user in rows:
        uid = int(payout.user_id)
        username = f"@{user.username}" if user.username else f"@{user.telegram_id}"
        kb_rows.append(
            [
                InlineKeyboardButton(
                    text=_pay_op_label(username),
                    callback_data=f"{CB_PAY_MARK}:{uid}:{page}",
                )
            ]
        )
    if total > LEDGER_PAGE_SIZE:
        nav: list[InlineKeyboardButton] = []
        if page > 0:
            nav.append(InlineKeyboardButton(text="<<", callback_data=f"{CB_PAY_LEDGER_PAGE}:{page - 1}"))
        nav.append(InlineKeyboardButton(text=f"{page + 1}/{max_page + 1}", callback_data=CB_NOOP))
        if page < max_page:
            nav.append(InlineKeyboardButton(text=">>", callback_data=f"{CB_PAY_LEDGER_PAGE}:{page + 1}"))
        kb_rows.append(nav)
    kb_rows.append(
        [
            InlineKeyboardButton(
                text="💳 Добавить USDT",
                callback_data=f"{CB_PAY_TOPUP}:{page}",
            )
        ]
    )
    kb_rows.append(
        [
            InlineKeyboardButton(text="История выплат", callback_data=f"{CB_PAY_HISTORY_PAGE}:0"),
            InlineKeyboardButton(text="Корзина", callback_data=f"{CB_PAY_TRASH_PAGE}:0"),
        ]
    )
    kb_rows.append(
        [
            InlineKeyboardButton(text="Управление PENDING", callback_data=f"{CB_PAY_PENDING_PAGE}:0"),
        ]
    )
    kb_rows.append([InlineKeyboardButton(text=REPLY_BTN_BACK, callback_data=CALLBACK_INLINE_BACK)])
    return text, InlineKeyboardMarkup(inline_keyboard=kb_rows)


async def _edit_payout_ledger_message(message: Message, session: AsyncSession, *, page: int) -> None:
    text, kb = await _payout_ledger_text_and_markup(session, page=page)
    await edit_message_text_safe(message, text, reply_markup=kb)


async def _payout_history_text_and_markup(session: AsyncSession, *, page: int) -> tuple[str, InlineKeyboardMarkup]:
    page = max(page, 0)
    rows, total = await BillingService(session=session).get_payouts_paginated(
        status=PayoutStatus.PAID,
        page=page,
        page_size=PAGE_SIZE,
    )
    max_page = max((max(total, 1) - 1) // PAGE_SIZE, 0)
    if page > max_page:
        page = max_page
        rows, total = await BillingService(session=session).get_payouts_paginated(
            status=PayoutStatus.PAID,
            page=page,
            page_size=PAGE_SIZE,
        )
    lines = ["📜 История выплат", ""]
    if not rows:
        lines.append("Пока пусто.")
    else:
        for payout, user in rows:
            username = f"@{user.username}" if user.username else f"@{user.telegram_id}"
            lines.append(f"- {payout.period_key} | {username} | {payout.amount} USDT")
    text = "\n".join(lines)
    nav: list[InlineKeyboardButton] = []
    if page > 0:
        nav.append(InlineKeyboardButton(text="⬅️", callback_data=f"{CB_PAY_HISTORY_PAGE}:{page - 1}"))
    nav.append(InlineKeyboardButton(text=f"{page + 1}/{max_page + 1}", callback_data=CB_NOOP))
    if page < max_page:
        nav.append(InlineKeyboardButton(text="➡️", callback_data=f"{CB_PAY_HISTORY_PAGE}:{page + 1}"))
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            nav,
            [InlineKeyboardButton(text="💰 К ведомости", callback_data=f"{CB_PAY_LEDGER_PAGE}:0")],
            [InlineKeyboardButton(text=REPLY_BTN_BACK, callback_data=CALLBACK_INLINE_BACK)],
        ]
    )
    return text, kb


async def _payout_trash_text_and_markup(session: AsyncSession, *, page: int) -> tuple[str, InlineKeyboardMarkup]:
    page = max(page, 0)
    rows, total = await BillingService(session=session).get_payouts_paginated(
        status=PayoutStatus.CANCELLED,
        page=page,
        page_size=PAGE_SIZE,
    )
    max_page = max((max(total, 1) - 1) // PAGE_SIZE, 0)
    if page > max_page:
        page = max_page
        rows, total = await BillingService(session=session).get_payouts_paginated(
            status=PayoutStatus.CANCELLED,
            page=page,
            page_size=PAGE_SIZE,
        )
    lines = ["🗑 Корзина выплат", ""]
    if not rows:
        lines.append("Пока пусто.")
    else:
        for payout, user in rows:
            username = f"@{user.username}" if user.username else f"@{user.telegram_id}"
            lines.append(f"- {payout.period_key} | {username} | {payout.amount} USDT")
    text = "\n".join(lines)
    nav: list[InlineKeyboardButton] = []
    if page > 0:
        nav.append(InlineKeyboardButton(text="⬅️", callback_data=f"{CB_PAY_TRASH_PAGE}:{page - 1}"))
    nav.append(InlineKeyboardButton(text=f"{page + 1}/{max_page + 1}", callback_data=CB_NOOP))
    if page < max_page:
        nav.append(InlineKeyboardButton(text="➡️", callback_data=f"{CB_PAY_TRASH_PAGE}:{page + 1}"))
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            nav,
            [InlineKeyboardButton(text="💰 К ведомости", callback_data=f"{CB_PAY_LEDGER_PAGE}:0")],
            [InlineKeyboardButton(text=REPLY_BTN_BACK, callback_data=CALLBACK_INLINE_BACK)],
        ]
    )
    return text, kb


async def _payout_pending_manage_text_and_markup(
    session: AsyncSession,
    *,
    page: int,
) -> tuple[str, InlineKeyboardMarkup]:
    page = max(page, 0)
    rows, total = await BillingService(session=session).get_payouts_paginated(
        status=PayoutStatus.PENDING,
        page=page,
        page_size=PAGE_SIZE,
    )
    max_page = max((max(total, 1) - 1) // PAGE_SIZE, 0)
    if page > max_page:
        page = max_page
        rows, total = await BillingService(session=session).get_payouts_paginated(
            status=PayoutStatus.PENDING,
            page=page,
            page_size=PAGE_SIZE,
        )

    lines = ["⚙️ Управление PENDING выплатами", ""]
    kb_rows: list[list[InlineKeyboardButton]] = []

    if not rows:
        lines.append("PENDING выплат нет.")
    else:
        for payout, user in rows:
            username = f"@{user.username}" if user.username else f"@{user.telegram_id}"
            lines.append(
                f"#{payout.id} | {payout.period_key} | {username} | {payout.accepted_count} шт. | {payout.amount} USDT"
            )
            kb_rows.append(
                [
                    InlineKeyboardButton(
                        text=f"🗑 Удалить #{payout.id}",
                        callback_data=f"{CB_PAY_PENDING_DELETE}:{payout.id}:{page}",
                    )
                ]
            )

    nav: list[InlineKeyboardButton] = []
    if page > 0:
        nav.append(InlineKeyboardButton(text="⬅️", callback_data=f"{CB_PAY_PENDING_PAGE}:{page - 1}"))
    nav.append(InlineKeyboardButton(text=f"{page + 1}/{max_page + 1}", callback_data=CB_NOOP))
    if page < max_page:
        nav.append(InlineKeyboardButton(text="➡️", callback_data=f"{CB_PAY_PENDING_PAGE}:{page + 1}"))
    kb_rows.append(nav)

    kb_rows.append([InlineKeyboardButton(text="💰 К ведомости", callback_data=f"{CB_PAY_LEDGER_PAGE}:0")])
    kb_rows.append([InlineKeyboardButton(text=REPLY_BTN_BACK, callback_data=CALLBACK_INLINE_BACK)])

    return "\n".join(lines), InlineKeyboardMarkup(inline_keyboard=kb_rows)


def _admin_panel_intro_text() -> str:
    return ADMIN_PANEL_HOME_TEXT


async def _fetch_admin_board_stats(session: AsyncSession) -> dict[str, int]:
    pending_stmt = select(func.count(Submission.id)).where(Submission.status == SubmissionStatus.PENDING)
    in_review_stmt = select(func.count(Submission.id)).where(Submission.status == SubmissionStatus.IN_REVIEW)
    approved_stmt = select(func.count(Submission.id)).where(Submission.status == SubmissionStatus.ACCEPTED)
    rejected_stmt = select(func.count(Submission.id)).where(
        Submission.status.in_(
            [
                SubmissionStatus.REJECTED,
                SubmissionStatus.BLOCKED,
                SubmissionStatus.NOT_A_SCAN,
            ]
        )
    )

    pending_count = int((await session.execute(pending_stmt)).scalar_one())
    in_review_count = int((await session.execute(in_review_stmt)).scalar_one())
    approved_count = int((await session.execute(approved_stmt)).scalar_one())
    rejected_count = int((await session.execute(rejected_stmt)).scalar_one())
    return {
        "pending_count": pending_count,
        "in_review_count": in_review_count,
        "approved_count": approved_count,
        "rejected_count": rejected_count,
    }


async def _render_admin_moderation_card(
    *,
    session: AsyncSession,
    submission: Submission,
) -> str:
    svc = SubmissionService(session=session)
    is_duplicate = await svc.has_phone_duplicate(
        submission_id=submission.id,
        phone=submission.description_text,
    )
    return GDPXRenderer().render_moderation_card(submission, is_duplicate=is_duplicate)


def _admin_section_inline_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text=REPLY_BTN_BACK, callback_data=CALLBACK_INLINE_BACK)],
        ]
    )


def _admin_inwork_inline_keyboard(items: list[Submission]) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    for item in items:
        phone = (item.description_text or "").strip() or "—"
        rows.append(
            [
                InlineKeyboardButton(
                    text=f"📱 {phone[:18]}",
                    callback_data=f"{CB_ADMIN_INWORK_OPEN}:{item.id}",
                )
            ]
        )
    rows.append([InlineKeyboardButton(text="🔍 Найти", callback_data=CB_ADMIN_INWORK_SEARCH)])
    rows.append([InlineKeyboardButton(text=REPLY_BTN_BACK, callback_data=CALLBACK_INLINE_BACK)])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _admin_payouts_inline_keyboard(rows_data: list[dict[str, int | str | Decimal]]) -> InlineKeyboardMarkup:
    rows: list[list[InlineKeyboardButton]] = []
    for row in rows_data:
        username = str(row.get("username", "—"))
        uid = int(row["user_id"])
        rows.append(
            [
                InlineKeyboardButton(
                    text=f"💳 Оплатить {username}",
                    callback_data=f"{CB_PAY_MARK}:{uid}:0",
                )
            ]
        )
    rows.append([InlineKeyboardButton(text=REPLY_BTN_BACK, callback_data=CALLBACK_INLINE_BACK)])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _admin_queue_lobby_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="🚀 Приступить к проверке", callback_data=CB_ADMIN_QUEUE_START)],
            [InlineKeyboardButton(text=REPLY_BTN_BACK, callback_data=CALLBACK_INLINE_BACK)],
        ]
    )


def _reply_matches_menu_label(expected: str):
    """Совпадение текста reply-кнопки с учётом регистра и пробелов."""

    def _check(t: str | None) -> bool:
        return match_admin_menu_canonical(t) == expected

    return _check


_ADMIN_FSM_STATES = (
    AdminBroadcastState.waiting_for_text,
    AdminSearchSimState.waiting_for_digits,
    AdminModerationForwardState.waiting_for_target,
    AdminModerationForwardState.waiting_for_confirm,
    AdminBatchPickState.waiting_for_submission_ids,
    AdminBatchPickState.waiting_for_action,
    AdminInReviewLookupState.waiting_for_query,
    AdminPayoutState.waiting_for_topup_amount,
)


@router.message(F.text == REPLY_BTN_BACK, StateFilter(*_ADMIN_FSM_STATES))
async def on_admin_fsm_step_back(
    message: Message,
    state: FSMContext,
    session: AsyncSession,
) -> None:
    """Шаг назад в админских сценариях или выход в главное меню."""

    if message.from_user is None:
        return
    if not await AdminService(session=session).is_admin(message.from_user.id):
        return

    st = await state.get_state()
    if st == AdminBroadcastState.waiting_for_text.state:
        await state.clear()
        await on_admin_panel(message, session)
        return
    if st == AdminSearchSimState.waiting_for_digits.state:
        await state.clear()
        await on_admin_panel(message, session)
        return
    if st == AdminModerationForwardState.waiting_for_target.state:
        await state.clear()
        pending_count = int(
            (
                await session.execute(
                    select(func.count(Submission.id)).where(Submission.status == SubmissionStatus.PENDING)
                )
            ).scalar_one()
        )
        text = GDPXRenderer().render_queue_lobby(pending_count=pending_count)
        await message.answer(text, reply_markup=_admin_queue_lobby_keyboard(), parse_mode="HTML")
        return
    if st == AdminModerationForwardState.waiting_for_confirm.state:
        await state.clear()
        pending_count = int(
            (
                await session.execute(
                    select(func.count(Submission.id)).where(Submission.status == SubmissionStatus.PENDING)
                )
            ).scalar_one()
        )
        text = GDPXRenderer().render_queue_lobby(pending_count=pending_count)
        await message.answer(text, reply_markup=_admin_queue_lobby_keyboard(), parse_mode="HTML")
        return
    if st == AdminBatchPickState.waiting_for_submission_ids.state:
        await state.clear()
        pending_count = int(
            (
                await session.execute(
                    select(func.count(Submission.id)).where(Submission.status == SubmissionStatus.PENDING)
                )
            ).scalar_one()
        )
        text = GDPXRenderer().render_queue_lobby(pending_count=pending_count)
        await message.answer(text, reply_markup=_admin_queue_lobby_keyboard(), parse_mode="HTML")
        return
    if st == AdminBatchPickState.waiting_for_action.state:
        await state.clear()
        pending_count = int(
            (
                await session.execute(
                    select(func.count(Submission.id)).where(Submission.status == SubmissionStatus.PENDING)
                )
            ).scalar_one()
        )
        text = GDPXRenderer().render_queue_lobby(pending_count=pending_count)
        await message.answer(text, reply_markup=_admin_queue_lobby_keyboard(), parse_mode="HTML")
        return
    if st == AdminInReviewLookupState.waiting_for_query.state:
        await state.clear()
        await on_in_work_hub(message, state, session)
        return
    if st == AdminPayoutState.waiting_for_topup_amount.state:
        await state.clear()
        await on_daily_report(message, state, session)
        return


@router.callback_query(F.data == CB_NOOP)
@router.callback_query(F.data == CB_NOOP)
async def on_noop(callback: CallbackQuery) -> None:
    await callback.answer()


async def on_in_work_hub(message: Message, state: FSMContext, session: AsyncSession) -> None:
    """Хаб «В работе»: компактная статистика и поиск по номеру."""

    if message.from_user is None:
        return
    if not await AdminService(session=session).is_admin(message.from_user.id):
        await message.answer("Недостаточно прав.")
        return

    admin_user = await UserService(session=session).get_by_telegram_id(message.from_user.id)
    if admin_user is None:
        await message.answer("Пользователь не найден в БД.")
        return

    await _admin_delete_prev_panel_message(message, state)
    mine = await SubmissionService(session=session).get_admin_active_submissions(admin_id=admin_user.id)
    await state.clear()
    total = len(mine)
    page = 0

    # Проверяем, chief_admin ли это
    is_chief = admin_user.role == "chief_admin"

    chunk = mine[:INWORK_PAGE_SIZE]
    text = GDPXRenderer().render_inwork_hub(chunk, is_chief=is_chief, index_offset=0)
    sent = await message.answer(
        text,
        reply_markup=_in_work_hub_keyboard(items=chunk, page=page, total=total),
        parse_mode="HTML",
    )
    await _admin_store_panel_message(state, sent)


@router.callback_query(F.data.startswith(f"{CB_ADMIN_INWORK_PAGE}:"))
async def on_in_work_page(callback: CallbackQuery, session: AsyncSession, state: FSMContext) -> None:
    if callback.from_user is None or callback.data is None:
        return
    if not await AdminService(session=session).is_admin(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    admin_user = await UserService(session=session).get_by_telegram_id(callback.from_user.id)
    if admin_user is None:
        await callback.answer("Пользователь не найден в БД", show_alert=True)
        return

    is_chief = admin_user.role == "chief_admin"
    page = max(int(callback.data.rsplit(":", 1)[-1]), 0)
    mine = await SubmissionService(session=session).get_admin_active_submissions(admin_id=admin_user.id)
    total = len(mine)
    if not mine:
        page = 0
        chunk: list = []
    else:
        max_page = max((total - 1) // INWORK_PAGE_SIZE, 0)
        page = min(page, max_page)
        chunk = mine[page * INWORK_PAGE_SIZE : page * INWORK_PAGE_SIZE + INWORK_PAGE_SIZE]

    base = page * INWORK_PAGE_SIZE
    text = GDPXRenderer().render_inwork_hub(chunk, is_chief=is_chief, index_offset=base)
    await callback.answer()
    if callback.message is not None:
        await edit_message_text_safe(
            callback.message,
            text,
            reply_markup=_in_work_hub_keyboard(items=chunk, page=page, total=total),
            parse_mode="HTML",
        )


@router.callback_query(F.data.startswith(f"{CB_ADMIN_INWORK_OPEN}:"))
async def on_in_work_open_submission(callback: CallbackQuery, session: AsyncSession) -> None:
    if callback.from_user is None or callback.data is None:
        return
    if not await AdminService(session=session).is_admin(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return

    admin_user = await UserService(session=session).get_by_telegram_id(callback.from_user.id)
    if admin_user is None:
        await callback.answer("Пользователь не найден в БД", show_alert=True)
        return

    try:
        submission_id = int(callback.data.rsplit(":", 1)[-1])
    except (TypeError, ValueError):
        await callback.answer("Некорректные данные кнопки.", show_alert=True)
        return

    svc = SubmissionService(session=session)
    item = await svc.get_submission_in_work_for_admin(
        submission_id=submission_id,
        admin_id=admin_user.id,
    )
    if item is None:
        await callback.answer("Эта заявка не в вашем списке «В работе».", show_alert=True)
        return

    cap = await _render_admin_moderation_card(session=session, submission=item)
    await callback.answer()
    if callback.message is not None:
        await message_answer_submission(
            callback.message,
            item,
            caption=cap,
            reply_markup=moderation_review_keyboard(submission_id=item.id),
            parse_mode="HTML",
        )


@router.callback_query(F.data == CB_ADMIN_INWORK_SEARCH)
async def on_in_work_search_start(callback: CallbackQuery, state: FSMContext, session: AsyncSession) -> None:
    if callback.from_user is None:
        return
    if not await AdminService(session=session).is_admin(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    await state.set_state(AdminInReviewLookupState.waiting_for_query)
    await callback.answer()
    if callback.message is not None:
        await edit_message_text_safe(
            callback.message,
            GDPXRenderer().render_inwork_search_prompt(),
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[
                    [InlineKeyboardButton(text=REPLY_BTN_BACK, callback_data=CALLBACK_INLINE_BACK)]
                ]
            ),
            parse_mode="HTML",
        )


@router.message(AdminInReviewLookupState.waiting_for_query, F.text)
async def on_in_work_search_query(message: Message, state: FSMContext, session: AsyncSession) -> None:
    if message.from_user is None or message.text is None:
        return
    if not await AdminService(session=session).is_admin(message.from_user.id):
        await state.clear()
        return

    query = message.text.strip()
    if not query:
        await message.answer("Нужно ввести номер.")
        return

    if query.startswith("+7") and len(query) == 12:
        where_clause = Submission.description_text == query
    else:
        digits = re.sub(r"\D", "", query)
        if len(digits) < 3:
            await message.answer("Укажи минимум 3 цифры или полный номер +7XXXXXXXXXX.")
            return
        where_clause = Submission.description_text.like(f"%{digits}")

    stmt = (
        select(Submission)
        .options(
            joinedload(Submission.category),
            joinedload(Submission.seller),
            joinedload(Submission.locked_by_admin),
        )
        .where(
            Submission.status == SubmissionStatus.IN_REVIEW,
            where_clause,
        )
        .order_by(Submission.assigned_at.desc().nullslast(), Submission.id.desc())
        .limit(10)
    )
    rows = list((await session.execute(stmt)).scalars().all())
    if not rows:
        await message.answer("В «В работе» ничего не найдено по этому номеру.")
        return

    for submission in rows:
        cap = await _render_admin_moderation_card(session=session, submission=submission)
        await message_answer_submission(
            message,
            submission,
            caption=cap,
            reply_markup=moderation_review_keyboard(submission_id=submission.id),
            parse_mode="HTML",
        )


@router.message(F.text.func(is_admin_main_menu_text), StateFilter(*_ADMIN_FSM_STATES))
async def on_admin_menu_interrupt_fsm(
    message: Message,
    state: FSMContext,
    session: AsyncSession,
) -> None:
    """Сбрасывает админский FSM при нажатии кнопки меню и выполняет выбранный раздел."""

    if message.from_user is None or message.text is None:
        return
    if not await AdminService(session=session).is_admin(message.from_user.id):
        return

    label = match_admin_menu_canonical(message.text)
    if label is None:
        return

    data = await state.get_data()
    old_panel = data.get(_ADMIN_LAST_PANEL_MSG_KEY)
    await state.clear()
    if old_panel is not None and message.chat is not None:
        try:
            await message.bot.delete_message(chat_id=message.chat.id, message_id=int(old_panel))
        except TelegramBadRequest:
            pass

    if label == "Очередь":
        await on_moderation_queue(message, session)
    elif label in {"В работе", "🏃 В работе"}:
        await on_in_work_hub(message, state, session)
    elif label == "Выплаты":
        await on_daily_report(message, state, session)
    elif label == "Рассылка":
        await on_broadcast_start(message, state, session)



@router.message(Command("admin"), StateFilter(*_ADMIN_FSM_STATES))
async def on_admin_command_interrupt_fsm(
    message: Message,
    state: FSMContext,
    session: AsyncSession,
) -> None:
    """Сбрасывает админский FSM и открывает inline-дашборд по /admin."""

    if message.from_user is None:
        return
    if not await AdminService(session=session).is_admin(message.from_user.id):
        return

    data = await state.get_data()
    old_panel = data.get(_ADMIN_LAST_PANEL_MSG_KEY)
    await state.clear()
    if old_panel is not None and message.chat is not None:
        try:
            await message.bot.delete_message(chat_id=message.chat.id, message_id=int(old_panel))
        except TelegramBadRequest:
            pass

    await on_admin_panel(message, session)


@router.message(F.text == BUTTON_ENTER_ADMIN_PANEL)
async def on_enter_admin_panel(message: Message, session: AsyncSession) -> None:
    """Открывает инлайн-панель админа (убирает reply-клавиатуру)."""

    if message.from_user is None:
        return
    admin_service = AdminService(session=session)
    if not await admin_service.is_admin(message.from_user.id):
        await message.answer("Недостаточно прав.")
        return
    # Убираем reply-клавиатуру с экрана пользователя
    await message.answer("⁠", reply_markup=ReplyKeyboardRemove())
    stats = await _fetch_admin_board_stats(session)
    stats["username"] = message.from_user.username or str(message.from_user.id)
    text = GDPXRenderer().render_admin_dashboard(stats)
    await message.answer(
        text,
        reply_markup=await build_admin_main_inline_keyboard(session, message.from_user.id),
        parse_mode="HTML",
    )


@router.message(F.text == BUTTON_EXIT_ADMIN_PANEL)
async def on_exit_admin_panel(message: Message, session: AsyncSession) -> None:
    """Возвращает обычное меню селлера + кнопку входа в админ-панель."""

    if message.from_user is None:
        return
    admin_service = AdminService(session=session)
    if not await admin_service.is_admin(message.from_user.id):
        await message.answer("Недостаточно прав.")
        return
    user = await UserService(session=session).get_by_telegram_id(message.from_user.id)
    if user is None:
        await message.answer("Пользователь не найден в БД.")
        return
    await message.answer(
        "Вы вышли из админ-панели. Профиль и сделки — в обычном меню ниже.",
        reply_markup=seller_main_menu_keyboard(language=user.language, role=user.role),
    )


@router.message(Command("admin"))
@router.message(Command("a"))
async def on_admin_panel(message: Message, session: AsyncSession) -> None:
    """Открывает главное меню админа."""

    if message.from_user is None:
        return

    admin_service = AdminService(session=session)
    if not await admin_service.is_admin(message.from_user.id):
        await message.answer("Недостаточно прав.")
        return

    await message.answer("\u2060", reply_markup=ReplyKeyboardRemove())
    stats = await _fetch_admin_board_stats(session)
    stats["username"] = message.from_user.username or str(message.from_user.id)
    text = GDPXRenderer().render_admin_dashboard(stats)
    await message.answer(
        text,
        reply_markup=await build_admin_main_inline_keyboard(session, message.from_user.id),
        parse_mode="HTML",
    )


@router.callback_query(F.data == CALLBACK_INLINE_BACK)
async def on_back_to_admin_menu(callback: CallbackQuery, state: FSMContext, session: AsyncSession) -> None:
    if callback.from_user is None:
        return
    if not await AdminService(session=session).is_admin(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return

    await state.clear()
    stats = await _fetch_admin_board_stats(session)
    stats["username"] = callback.from_user.username or str(callback.from_user.id)
    text = GDPXRenderer().render_admin_dashboard(stats)
    await callback.answer()
    if callback.message is not None:
        try:
            await callback.message.edit_text(
                text=text,
                reply_markup=await build_admin_main_inline_keyboard(session, callback.from_user.id),
                parse_mode="HTML",
            )
        except TelegramBadRequest:
            await callback.message.answer(
                text=text,
                reply_markup=await build_admin_main_inline_keyboard(session, callback.from_user.id),
                parse_mode="HTML",
            )


@router.callback_query(F.data == CB_ADMIN_QUEUE)
async def on_admin_inline_queue(callback: CallbackQuery, session: AsyncSession) -> None:
    if callback.from_user is None:
        return
    if not await AdminService(session=session).is_admin(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    pending_count = int(
        (
            await session.execute(
                select(func.count(Submission.id)).where(Submission.status == SubmissionStatus.PENDING)
            )
        ).scalar_one()
    )
    text = GDPXRenderer().render_queue_lobby(pending_count=pending_count)
    await callback.answer()
    if callback.message is not None:
        try:
            await callback.message.edit_text(
                text=text,
                reply_markup=_admin_queue_lobby_keyboard(),
                parse_mode="HTML",
            )
        except TelegramBadRequest:
            pass


@router.callback_query(F.data == CB_ADMIN_INWORK_HUB)
async def on_admin_inline_inwork(callback: CallbackQuery, session: AsyncSession, state: FSMContext) -> None:
    if callback.from_user is None:
        return
    if not await AdminService(session=session).is_admin(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return

    admin_user = await UserService(session=session).get_by_telegram_id(callback.from_user.id)
    if admin_user is None:
        await callback.answer("Пользователь не найден в БД", show_alert=True)
        return

    mine = await SubmissionService(session=session).get_admin_active_submissions(admin_id=admin_user.id)
    await state.clear()
    total = len(mine)
    is_chief = admin_user.role == "chief_admin"

    chunk = mine[:INWORK_PAGE_SIZE]
    text = GDPXRenderer().render_inwork_hub(chunk, is_chief=is_chief, index_offset=0)
    markup = _in_work_hub_keyboard(items=chunk, page=0, total=total)

    await callback.answer()
    if callback.message is not None:
        await edit_message_text_safe(callback.message, text, reply_markup=markup, parse_mode="HTML")


@router.callback_query(F.data == CB_ADMIN_PAYOUTS)
async def on_admin_inline_payouts(callback: CallbackQuery, session: AsyncSession, state: FSMContext) -> None:
    if callback.from_user is None:
        return
    if not await AdminService(session=session).can_manage_payouts(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    await callback.answer()
    if callback.message is not None:
        await on_daily_report(callback.message, state, session, _caller_id=callback.from_user.id)


@router.callback_query(F.data == CB_ADMIN_STATS_MONTH)
async def on_admin_inline_stats_month(callback: CallbackQuery, session: AsyncSession) -> None:
    if callback.from_user is None:
        return
    if not await AdminService(session=session).can_manage_payouts(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return

    now = datetime.now(timezone.utc)
    year, month = now.year, now.month
    rows = await AdminStatsService(session=session).daily_sim_stats_for_month(year=year, month=month)
    text = _month_stats_text(rows, year=year, month=month)

    await callback.answer()
    if callback.message is not None:
        await edit_message_text_safe(
            callback.message,
            text,
            reply_markup=_month_stats_keyboard(year=year, month=month),
            parse_mode="HTML",
        )


@router.callback_query(F.data.startswith(f"{CB_ADMIN_STATS_EXPORT_MONTH}:"))
async def on_admin_inline_stats_export_month(callback: CallbackQuery, session: AsyncSession) -> None:
    if callback.from_user is None or callback.data is None:
        return
    if not await AdminService(session=session).can_manage_payouts(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return

    parts = callback.data.split(":")
    if len(parts) != 4:
        await callback.answer("Некорректные параметры экспорта", show_alert=True)
        return

    try:
        year = int(parts[2])
        month = int(parts[3])
    except ValueError:
        await callback.answer("Некорректный период", show_alert=True)
        return

    rows = await AdminStatsService(session=session).daily_sim_stats_for_month(year=year, month=month)
    payload = _month_stats_workbook_bytes(rows, year=year, month=month)

    await callback.answer("Формирую Excel...")
    if callback.message is not None:
        await callback.message.answer_document(
            document=(f"sim_stats_{year}_{month:02d}.xlsx", payload),
            caption=f"📊 Отчёт по SIM за {month:02d}.{year}",
        )


@router.callback_query(F.data == CB_ADMIN_STATS_RESET)
async def on_admin_stats_reset_ask(callback: CallbackQuery, session: AsyncSession) -> None:
    """Запрос подтверждения обнуления статистики (только chief_admin)."""
    if callback.from_user is None:
        return
    if not await AdminService(session=session).can_manage_payouts(callback.from_user.id):
        await callback.answer("Только для главных админов", show_alert=True)
        return
    await callback.answer()
    if callback.message is not None:
        await edit_message_text_safe(
            callback.message,
            "⚠️ <b>Обнулить статистику?</b>\n\n"
            "Все счётчики в разделе «Статистика» станут 0.\n"
            "Данные в БД не удаляются — просто сдвигается точка отсчёта.",
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        InlineKeyboardButton(
                            text="✅ Да, обнулить",
                            callback_data=CB_ADMIN_STATS_RESET_CONFIRM,
                        ),
                        InlineKeyboardButton(
                            text="❌ Отмена",
                            callback_data=CB_ADMIN_STATS_MONTH,
                        ),
                    ]
                ]
            ),
            parse_mode="HTML",
        )


@router.callback_query(F.data == CB_ADMIN_STATS_RESET_CONFIRM)
async def on_admin_stats_reset_confirm(callback: CallbackQuery, session: AsyncSession) -> None:
    """Выполняет обнуление: сохраняет текущее время как точку отсчёта."""
    if callback.from_user is None:
        return
    if not await AdminService(session=session).can_manage_payouts(callback.from_user.id):
        await callback.answer("Только для главных админов", show_alert=True)
        return

    from src.core.stats_epoch import set_stats_epoch
    epoch = set_stats_epoch()

    await AdminAuditService(session=session).log(
        admin_id=(await UserService(session=session).get_by_telegram_id(callback.from_user.id)).id,
        action="stats_reset",
        target_type="system",
        details=f"Stats epoch set to {epoch.isoformat()}",
    )

    now = datetime.now(timezone.utc)
    year, month = now.year, now.month
    rows = await AdminStatsService(session=session).daily_sim_stats_for_month(year=year, month=month)
    text = _month_stats_text(rows, year=year, month=month)

    await callback.answer("Статистика обнулена ✅")
    if callback.message is not None:
        await edit_message_text_safe(
            callback.message,
            text,
            reply_markup=_month_stats_keyboard(year=year, month=month),
            parse_mode="HTML",
        )


@router.callback_query(F.data == CB_ADMIN_DASHBOARD_RESET)
async def on_admin_dashboard_reset_ask(callback: CallbackQuery, session: AsyncSession) -> None:
    """Запрос подтверждения персонального сброса счётчиков дашборда."""
    if callback.from_user is None:
        return
    if not await AdminService(session=session).is_admin(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    await callback.answer()
    if callback.message is not None:
        await edit_message_text_safe(
            callback.message,
            "🔄 <b>Сбросить личные счётчики?</b>\n\n"
            "Счётчики «Принято всего» и «Брак всего» на твоём дашборде\n"
            "начнут считаться с нуля от текущего момента.\n\n"
            "Данные других админов не затрагиваются.",
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        InlineKeyboardButton(
                            text="✅ Да, обнулить",
                            callback_data=CB_ADMIN_DASHBOARD_RESET_CONFIRM,
                        ),
                        InlineKeyboardButton(
                            text="❌ Отмена",
                            callback_data=CALLBACK_INLINE_BACK,
                        ),
                    ]
                ]
            ),
            parse_mode="HTML",
        )


@router.callback_query(F.data == CB_ADMIN_DASHBOARD_RESET_CONFIRM)
async def on_admin_dashboard_reset_confirm(callback: CallbackQuery, session: AsyncSession) -> None:
    """Сохраняет персональную точку сброса, показывает обновлённый дашборд."""
    if callback.from_user is None:
        return
    if not await AdminService(session=session).is_admin(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return

    from src.core.personal_epoch import set_personal_epoch
    from src.utils.admin_keyboard import send_admin_dashboard
    tg_id = callback.from_user.id
    set_personal_epoch(tg_id)

    await callback.answer("Счётчики сброшены ✅")
    if callback.message is not None:
        await send_admin_dashboard(callback.message, session, tg_id)


@router.callback_query(F.data == CB_ADMIN_BROADCAST)
async def on_admin_inline_broadcast(callback: CallbackQuery, state: FSMContext, session: AsyncSession) -> None:
    """Запускает рассылку из инлайн-дашборда."""

    if callback.from_user is None:
        return
    if not await AdminService(session=session).is_admin(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    await state.set_state(AdminBroadcastState.waiting_for_text)
    await callback.answer()
    if callback.message is not None:
        await edit_message_text_safe(
            callback.message,
            f"📡 <b>РАССЫЛКА</b>\n\nОтправь текст рассылки одним сообщением.\n\n{HINT_BROADCAST}",
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[
                    [InlineKeyboardButton(text=REPLY_BTN_BACK, callback_data=CALLBACK_INLINE_BACK)]
                ]
            ),
            parse_mode="HTML",
        )


@router.callback_query(F.data == CB_ADMIN_QUEUE_START)
async def on_admin_queue_start(callback: CallbackQuery, session: AsyncSession) -> None:
    if callback.from_user is None:
        return
    if not await AdminService(session=session).is_admin(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    await callback.answer()
    if callback.message is not None:
        try:
            await callback.message.delete()
        except TelegramBadRequest:
            pass
        await on_moderation_queue(callback.message, session, _caller_id=callback.from_user.id)


async def on_broadcast_start(message: Message, state: FSMContext, session: AsyncSession) -> None:
    """Запускает массовую рассылку всем активным пользователям."""

    if message.from_user is None:
        return
    if not await AdminService(session=session).is_admin(message.from_user.id):
        await message.answer("Недостаточно прав.")
        return
    await state.set_state(AdminBroadcastState.waiting_for_text)
    await message.answer(f"Отправь текст рассылки одним сообщением.\n\n{HINT_BROADCAST}")


@router.message(AdminBroadcastState.waiting_for_text)
async def on_broadcast_send(
    message: Message,
    state: FSMContext,
    session: AsyncSession,
    bot: "Bot",
) -> None:
    """Отправляет массовую рассылку и показывает статистику доставки."""

    if message.from_user is None or message.text is None:
        return
    if not await AdminService(session=session).is_admin(message.from_user.id):
        await state.clear()
        await message.answer("Недостаточно прав.")
        return

    admin_user = await UserService(session=session).get_by_telegram_id(message.from_user.id)
    body = (message.text or "").strip()
    if not body:
        await message.answer("Текст рассылки не может быть пустым. Отправь непустой текст.")
        return

    recipients = await UserService(session=session).get_all_active_users()
    delivered = 0
    failed = 0
    for user in recipients:
        try:
            await bot.send_message(chat_id=user.telegram_id, text=body)
            delivered += 1
        except TelegramAPIError:
            failed += 1

    await state.clear()
    await message.answer(
        f"Рассылка завершена.\nУспешно: {delivered}\nОшибок: {failed}",
    )
    if admin_user is not None:
        await AdminAuditService(session=session).log(
            admin_id=admin_user.id,
            action="broadcast",
            target_type="users",
            details=f"delivered={delivered},failed={failed}",
        )


@router.message(Command("daily_report"))
async def on_daily_report(
    message: Message, state: FSMContext, session: AsyncSession, *, _caller_id: int | None = None
) -> None:
    """Показывает итоговую ведомость к выплате (одно сообщение)."""

    tid = _caller_id or (message.from_user.id if message.from_user else None)
    if tid is None:
        return

    admin_service = AdminService(session=session)
    if not await admin_service.can_manage_payouts(tid):
        await message.answer("Недостаточно прав.")
        return

    text, kb = await _payout_ledger_text_and_markup(session, page=0)
    body = f"{HINT_PAYOUTS}\n\n{text}"
    data = await state.get_data()
    mid = data.get(_ADMIN_LAST_PANEL_MSG_KEY)
    if mid is not None and message.chat is not None:
        try:
            await message.bot.edit_message_text(
                chat_id=message.chat.id,
                message_id=int(mid),
                text=body,
                reply_markup=kb,
            )
            return
        except TelegramBadRequest:
            pass

    await _admin_delete_prev_panel_message(message, state)
    sent = await message.answer(body, reply_markup=kb)
    await _admin_store_panel_message(state, sent)


@router.message(Command("s"))
async def on_search_submission(message: Message, session: AsyncSession) -> None:
    """Ищет товары в работе/истории по номеру или последним цифрам."""

    if message.from_user is None or message.text is None:
        return
    if not await AdminService(session=session).is_admin(message.from_user.id):
        await message.answer("Недостаточно прав.")
        return

    raw_query = message.text.replace("/s", "", 1).strip()
    if not raw_query:
        await message.answer("Формат: /s 1234 или /s +79999999999")
        return

    digits = re.sub(r"\D", "", raw_query)
    if not PHONE_QUERY_PATTERN.fullmatch(raw_query) and len(digits) < 3:
        await message.answer("Укажи минимум 3 последние цифры или полный номер.")
        return

    rows, total = await SubmissionService(session=session).search_by_phone_paginated(
        query=raw_query,
        page=0,
        page_size=PAGE_SIZE,
    )
    if not rows:
        await message.answer("Ничего не найдено по этому запросу.")
        return

    for submission, seller in rows:
        cap = await _render_admin_moderation_card(session=session, submission=submission)
        await message_answer_submission(
            message,
            submission,
            caption=cap,
            reply_markup=search_report_keyboard(submission_id=submission.id, seller_user_id=submission.user_id),
            parse_mode="HTML",
        )
    await message.answer(
        "Навигация поиска:",
        reply_markup=pagination_keyboard(
            CB_ADMIN_SEARCH_PAGE,
            page=0,
            total=total,
            page_size=PAGE_SIZE,
            query=raw_query,
        ),
    )


@router.callback_query(F.data.startswith(f"{CB_ADMIN_SEARCH_PAGE}:"))
async def on_search_page(callback: CallbackQuery, session: AsyncSession) -> None:
    if callback.from_user is None or callback.data is None:
        return
    if not await AdminService(session=session).is_admin(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    _, _, page_raw, query = callback.data.split(":", 3)
    page = max(int(page_raw), 0)
    rows, total = await SubmissionService(session=session).search_by_phone_paginated(
        query=query,
        page=page,
        page_size=PAGE_SIZE,
    )
    if callback.message is not None:
        for submission, seller in rows:
            cap = await _render_admin_moderation_card(session=session, submission=submission)
            await message_answer_submission(
                callback.message,
                submission,
                caption=cap,
                reply_markup=search_report_keyboard(submission_id=submission.id, seller_user_id=submission.user_id),
                parse_mode="HTML",
            )
        await callback.message.answer(
            "Навигация поиска:",
            reply_markup=pagination_keyboard(
                CB_ADMIN_SEARCH_PAGE,
                page=page,
                total=total,
                page_size=PAGE_SIZE,
                query=query,
            ),
        )
    await callback.answer()


@router.callback_query(F.data.startswith(f"{CB_ADMIN_RESTRICT}:"))
async def on_restrict_user(callback: CallbackQuery, session: AsyncSession) -> None:
    if callback.from_user is None or callback.data is None:
        return
    admin = await UserService(session=session).get_by_telegram_id(callback.from_user.id)
    if admin is None or not await AdminService(session=session).is_admin(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    user_id = int(callback.data.split(":")[2])
    target = await UserService(session=session).set_restricted(user_id=user_id, value=True)
    if target is None:
        await callback.answer("Пользователь не найден", show_alert=True)
        return
    await AdminAuditService(session=session).log(
        admin_id=admin.id,
        action="set_restricted",
        target_type="user",
        target_id=user_id,
        details="manual from admin report",
    )
    await callback.answer("Ограничение включено")


@router.callback_query(F.data.startswith(f"{CB_ADMIN_UNRESTRICT}:"))
async def on_unrestrict_user(callback: CallbackQuery, session: AsyncSession) -> None:
    if callback.from_user is None or callback.data is None:
        return
    admin = await UserService(session=session).get_by_telegram_id(callback.from_user.id)
    if admin is None or not await AdminService(session=session).is_admin(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    user_id = int(callback.data.split(":")[2])
    target = await UserService(session=session).set_restricted(user_id=user_id, value=False)
    if target is None:
        await callback.answer("Пользователь не найден", show_alert=True)
        return
    await AdminAuditService(session=session).log(
        admin_id=admin.id,
        action="unset_restricted",
        target_type="user",
        target_id=user_id,
        details="manual from admin report",
    )
    await callback.answer("Ограничение снято")


@router.message(Command("export_report"))
async def on_export_report(message: Message, session: AsyncSession) -> None:
    """Экспортирует отчёт выплат в CSV/XLSX."""

    if message.from_user is None or message.text is None:
        return
    if not await AdminService(session=session).is_admin(message.from_user.id):
        await message.answer("Недостаточно прав.")
        return
    fmt = message.text.replace("/export_report", "", 1).strip().lower() or "csv"
    rows = await BillingService(session=session).get_daily_report_rows()
    if not rows:
        await message.answer("Нет данных для экспорта.")
        return

    if fmt == "xlsx":
        try:
            from openpyxl import Workbook
        except Exception:
            await message.answer("Для XLSX установи зависимость openpyxl.")
            return
        wb = Workbook()
        ws = wb.active
        ws.append(["username", "accepted_count", "to_pay"])
        for row in rows:
            ws.append([str(row["username"]), int(row["accepted_count"]), str(row["to_pay"])])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        await message.answer_document(
            document=("daily_report.xlsx", buf.read()),
            caption="Экспорт XLSX готов.",
        )
        return

    sio = StringIO()
    writer = csv.writer(sio)
    writer.writerow(["username", "accepted_count", "to_pay"])
    for row in rows:
        writer.writerow([row["username"], row["accepted_count"], row["to_pay"]])
    await message.answer_document(
        document=("daily_report.csv", sio.getvalue().encode("utf-8")),
        caption="CSV-файл подготовлен.",
    )


@router.callback_query(F.data.startswith(f"{CB_PAY_LEDGER_PAGE}:"))
async def on_payout_ledger_page(callback: CallbackQuery, session: AsyncSession) -> None:
    if callback.from_user is None or callback.data is None:
        return
    if not await AdminService(session=session).can_manage_payouts(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    page = max(int(callback.data.rsplit(":", 1)[-1]), 0)
    await callback.answer()
    if callback.message is not None:
        await _edit_payout_ledger_message(callback.message, session, page=page)


def _parse_pay_topup_page(callback_data: str) -> int:
    parts = callback_data.split(":")
    if len(parts) < 3:
        return 0
    try:
        return max(int(parts[2]), 0)
    except ValueError:
        return 0


@router.callback_query(
    lambda c: c.data is not None and c.data.startswith(f"{CB_PAY_TOPUP}:") and len(c.data.split(":")) == 3
)
async def on_payout_topup_open(callback: CallbackQuery, session: AsyncSession, state: FSMContext) -> None:
    """Открывает ввод суммы пополнения CryptoPay из экрана ведомости."""

    if callback.from_user is None or callback.data is None:
        return
    if not await AdminService(session=session).can_manage_payouts(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return

    ledger_page = _parse_pay_topup_page(callback.data)
    await state.set_state(AdminPayoutState.waiting_for_topup_amount)
    await state.update_data(payout_topup_ledger_page=ledger_page)

    await callback.answer()
    if callback.message is not None:
        await edit_message_text_safe(
            callback.message,
            (
                "💳 <b>ПОПОЛНЕНИЕ APP БАЛАНСА</b>\n\n"
                "Введите сумму пополнения в USDT одним сообщением.\n"
                "Например: <code>1000</code> или <code>1000.50</code>."
            ),
            reply_markup=InlineKeyboardMarkup(
                inline_keyboard=[
                    [InlineKeyboardButton(text="⬅️ К ведомости", callback_data=f"{CB_PAY_LEDGER_PAGE}:{ledger_page}")],
                    [InlineKeyboardButton(text=REPLY_BTN_BACK, callback_data=CALLBACK_INLINE_BACK)],
                ]
            ),
            parse_mode="HTML",
        )


@router.message(AdminPayoutState.waiting_for_topup_amount, F.text)
async def on_payout_topup_amount_entered(message: Message, session: AsyncSession, state: FSMContext) -> None:
    """Создаёт invoice на введённую сумму для пополнения app-баланса CryptoPay."""

    if message.from_user is None:
        return
    if not await AdminService(session=session).can_manage_payouts(message.from_user.id):
        await state.clear()
        await message.answer("Недостаточно прав.")
        return

    raw = (message.text or "").strip().replace(",", ".")
    try:
        amount = Decimal(raw)
    except (InvalidOperation, TypeError):
        await message.answer("Некорректная сумма. Введите число, например: 1000 или 1000.50")
        return

    if amount <= Decimal("0"):
        await message.answer("Сумма должна быть больше 0.")
        return

    data = await state.get_data()
    ledger_page = int(data.get("payout_topup_ledger_page", 0))

    try:
        invoice = await CryptoBotService().create_topup_invoice(
            amount=amount,
            description=f"Manual top-up by chief admin {message.from_user.id}",
        )
    except RuntimeError as exc:
        await message.answer(f"Не удалось создать invoice: {exc}")
        return

    await state.clear()
    await message.answer(
        (
            "✅ <b>Invoice для пополнения создан</b>\n\n"
            f"<b>Сумма:</b> <code>{amount} USDT</code>\n"
            "\n"
            "После оплаты нажмите «✅ Я оплатил», бот проверит статус и вернёт в «Выплаты»."
        ),
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text="✅ Я оплатил",
                        callback_data=f"{CB_PAY_TOPUP_CHECK}:{invoice.invoice_id}:{ledger_page}:{amount}",
                    )
                ],
                [InlineKeyboardButton(text="Открыть", url=invoice.invoice_url)],
            ]
        ),
    )


@router.callback_query(lambda c: c.data is not None and c.data.startswith(f"{CB_PAY_TOPUP_CHECK}:"))
async def on_payout_topup_check_paid(callback: CallbackQuery, session: AsyncSession) -> None:
    """Проверяет статус topup-invoice и при оплате возвращает в раздел «Выплаты»."""

    if callback.from_user is None or callback.data is None:
        return
    if not await AdminService(session=session).can_manage_payouts(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return

    parts = callback.data.split(":", 4)
    if len(parts) < 5:
        await callback.answer("Некорректные данные invoice", show_alert=True)
        return
    try:
        invoice_id = int(parts[2])
        ledger_page = max(int(parts[3]), 0)
    except ValueError:
        await callback.answer("Некорректный invoice id", show_alert=True)
        return
    amount_label = parts[4]

    try:
        status = await CryptoBotService().get_invoice_status(invoice_id)
    except RuntimeError as exc:
        await callback.answer(f"Ошибка проверки invoice: {exc}", show_alert=True)
        return

    if status.status != "paid":
        await callback.answer("Invoice ещё не оплачен", show_alert=True)
        return

    await callback.answer(f"✅ APP пополнен на {amount_label} USDT", show_alert=True)
    if callback.message is not None:
        await _edit_payout_ledger_message(callback.message, session, page=ledger_page)


@router.callback_query(lambda c: c.data is not None and c.data.startswith(f"{CB_PAY_MARK}:"))
async def on_mark_paid(callback: CallbackQuery, session: AsyncSession, state: FSMContext) -> None:
    """Запрашивает подтверждение выплаты с показом статистики (первый уровень)."""

    if callback.from_user is None or callback.data is None:
        return

    admin_service = AdminService(session=session)
    if not await admin_service.can_manage_payouts(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return

    user_id, ledger_page = _parse_pay_uid_page(callback.data)
    if user_id <= 0:
        await callback.answer("Некорректные данные", show_alert=True)
        return
    user = await session.get(User, user_id)
    if user is None:
        await callback.answer("Пользователь не найден", show_alert=True)
        return

    # Получаем PENDING payouts для этого пользователя
    pending_stmt = (
        select(Payout)
        .where(
            Payout.user_id == user_id,
            Payout.status == PayoutStatus.PENDING,
        )
        .order_by(Payout.created_at.asc())
    )
    pending_payouts = list((await session.execute(pending_stmt)).scalars().all())
    
    if not pending_payouts:
        await callback.answer("Нет ожидающих выплат для этого пользователя", show_alert=True)
        return

    # Считаем статистику
    total_amount = Decimal("0.00")
    total_accepted_count = 0
    
    for payout in pending_payouts:
        total_amount += payout.amount
        total_accepted_count += payout.accepted_count

    # Получаем статистику отклонений за день
    day_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    
    rejected_stmt = select(func.count(Submission.id)).where(
        Submission.user_id == user_id,
        Submission.status.in_([
            SubmissionStatus.REJECTED,
            SubmissionStatus.BLOCKED,
            SubmissionStatus.NOT_A_SCAN,
        ]),
        Submission.reviewed_at >= day_start,
    )
    rejected_count = int((await session.execute(rejected_stmt)).scalar_one())

    # Сохраняем данные в FSM для следующего экрана подтверждения
    await state.set_state(AdminPayoutState.waiting_for_payout_confirm)
    await state.update_data(
        payout_user_id=user_id,
        payout_total_amount=str(total_amount),
        payout_accepted_count=total_accepted_count,
        payout_rejected_count=rejected_count,
        payout_username=f"@{user.username}" if user.username else f"@{user.telegram_id}",
        payout_ledger_page=ledger_page,
    )

    await callback.answer()
    if callback.message is not None:
        username = f"@{user.username}" if user.username else f"@{user.telegram_id}"
        stats_text = (
            f"💰 <b>ПОДТВЕРЖДЕНИЕ ВЫПЛАТЫ (Шаг 1)</b>\n\n"
            f"<b>Продавец:</b> {username}\n"
            f"<b>Период:</b> сегодня (UTC)\n\n"
            f"<b>📊 СТАТИСТИКА:</b>\n"
            f"✅ Зачёт (принято): {total_accepted_count} шт.\n"
            f"❌ Не зачёт: {rejected_count} шт.\n\n"
            f"<b>💵 СУММА К ВЫПЛАТЕ:</b> {total_amount} USDT\n\n"
            f"<i>Внимание: За каждой карточкой закреплена своя цена.</i>\n\n"
            f"<b>➡️ Нажмите «Подтвердить» для проверки данных...</b>"
        )
        await edit_message_text_safe(
            callback.message,
            stats_text,
            reply_markup=payout_confirm_keyboard(user_id=user_id, ledger_page=ledger_page),
            parse_mode="HTML",
        )


@router.callback_query(lambda c: c.data is not None and c.data.startswith(f"{CB_PAY_CANCEL}:"))
async def on_mark_paid_cancel(callback: CallbackQuery, session: AsyncSession, state: FSMContext) -> None:
    """Отменяет подтверждение выплаты и возвращает ведомость."""

    if callback.data is None:
        return
    await state.clear()
    user_id, ledger_page = _parse_pay_uid_page(callback.data)
    await callback.answer("Оплата отменена")
    if callback.message is not None:
        await _edit_payout_ledger_message(callback.message, session, page=ledger_page)


@router.callback_query(
    StateFilter(AdminPayoutState.waiting_for_payout_confirm),
    lambda c: c.data is not None and c.data.startswith(f"{CB_PAY_CONFIRM}:")
)
async def on_payout_confirmation(callback: CallbackQuery, session: AsyncSession, state: FSMContext) -> None:
    """Показывает финальное подтверждение перед отправкой чека в CryptoBot (шаг 2)."""

    if callback.from_user is None or callback.data is None:
        return

    data = await state.get_data()
    total_amount = data.get("payout_total_amount", "0.00")
    username = data.get("payout_username", "Unknown")
    try:
        payout_amount = Decimal(str(total_amount))
    except (InvalidOperation, TypeError):
        payout_amount = Decimal("0")

    available_usdt: Decimal | None = None
    balance_error: str | None = None
    try:
        available_usdt = await CryptoBotService().get_available_balance(asset_code="USDT")
    except RuntimeError as exc:
        balance_error = str(exc)
    
    if callback.message is not None:
        if balance_error is not None:
            balance_line = f"⚠️ <b>Баланс CryptoPay:</b> не удалось получить ({escape(balance_error)})"
            warning_line = "⚠️ <i>Проверьте токен/доступность CryptoPay перед отправкой.</i>"
        else:
            assert available_usdt is not None
            balance_line = f"💳 <b>Доступно в CryptoPay:</b> <code>{available_usdt} USDT</code>"
            warning_line = (
                "⚠️ <i>Недостаточно средств. Сначала пополните баланс через invoice.</i>"
                if available_usdt < payout_amount
                else "✅ <i>Средств достаточно для отправки чека.</i>"
            )

        final_text = (
            f"🔐 <b>ФИНАЛЬНОЕ ПОДТВЕРЖДЕНИЕ (Шаг 2)</b>\n\n"
            f"<b>Вы действительно хотите отправить чек?</b>\n\n"
            f"<b>Сумма:</b> <code>{total_amount} USDT</code>\n"
            f"<b>Получатель:</b> {username}\n"
            f"{balance_line}\n\n"
            f"{warning_line}\n"
            f"⚠️ <i>После отправки чека операцию нельзя будет отменить!</i>"
        )
        
        user_id = data.get("payout_user_id")
        ledger_page = data.get("payout_ledger_page", 0)
        
        await callback.answer()
        await edit_message_text_safe(
            callback.message,
            final_text,
            reply_markup=payout_final_confirm_keyboard(user_id=user_id, ledger_page=ledger_page),
            parse_mode="HTML",
        )


@router.callback_query(
    StateFilter(AdminPayoutState.waiting_for_payout_confirm),
    lambda c: c.data is not None and c.data.startswith(f"{CB_PAY_TOPUP}:")
)
async def on_create_topup_invoice(callback: CallbackQuery, state: FSMContext) -> None:
    """Создаёт invoice для пополнения баланса CryptoPay и показывает ссылку на оплату."""

    if callback.from_user is None or callback.data is None or callback.message is None:
        return

    data = await state.get_data()
    total_amount = data.get("payout_total_amount", "0.00")
    username = data.get("payout_username", "Unknown")
    user_id = data.get("payout_user_id")
    ledger_page = int(data.get("payout_ledger_page", 0))

    if user_id is None:
        await callback.answer("Данные сессии потеряны", show_alert=True)
        return

    try:
        invoice_amount = Decimal(str(total_amount))
    except (InvalidOperation, TypeError):
        invoice_amount = Decimal("0")
    if invoice_amount <= Decimal("0"):
        invoice_amount = Decimal("1")

    try:
        invoice = await CryptoBotService().create_topup_invoice(
            amount=invoice_amount,
            description=f"Top-up for payout {username}",
        )
    except RuntimeError as exc:
        await callback.answer("Не удалось создать invoice", show_alert=True)
        await edit_message_text_safe(
            callback.message,
            f"⚠️ Ошибка создания invoice:\n<code>{escape(str(exc))}</code>",
            reply_markup=payout_final_confirm_keyboard(user_id=user_id, ledger_page=ledger_page),
            parse_mode="HTML",
        )
        return

    await callback.answer("Invoice создан")
    await edit_message_text_safe(
        callback.message,
        (
            "💳 <b>Invoice для пополнения создан</b>\n\n"
            f"<b>Сумма:</b> <code>{invoice_amount} USDT</code>\n"
            "\n"
            "После оплаты нажмите «Точно отправить чек»."
        ),
        reply_markup=payout_final_confirm_keyboard(user_id=user_id, ledger_page=ledger_page),
        parse_mode="HTML",
    )


@router.callback_query(lambda c: c.data is not None and c.data.startswith(f"{CB_PAY_TRASH}:"))
async def on_mark_trash(callback: CallbackQuery, session: AsyncSession) -> None:
    """Отправляет выплату в корзину (cancelled)."""

    if callback.from_user is None or callback.data is None:
        return
    admin_service = AdminService(session=session)
    if not await admin_service.can_manage_payouts(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    admin_user = await UserService(session=session).get_by_telegram_id(callback.from_user.id)
    if admin_user is None:
        await callback.answer("Админ не найден в БД", show_alert=True)
        return
    user_id, ledger_page = _parse_pay_uid_page(callback.data)
    payout = await BillingService(session=session).cancel_user_payout(
        user_id=user_id,
        cancelled_by_admin_id=admin_user.id,
    )
    if payout is None:
        await callback.answer("Баланс уже пустой или выплата обработана", show_alert=True)
        return
    await callback.answer("Перемещено в корзину")
    await AdminAuditService(session=session).log(
        admin_id=admin_user.id,
        action="cancel_payout",
        target_type="user",
        target_id=user_id,
        details=f"amount={payout.amount}",
    )
    if callback.message is not None:
        await _edit_payout_ledger_message(callback.message, session, page=ledger_page)


@router.callback_query(
    StateFilter(AdminPayoutState.waiting_for_payout_confirm),
    lambda c: c.data is not None and c.data.startswith(f"{CB_PAY_FINAL_CONFIRM}:")
)
async def on_mark_paid_final(callback: CallbackQuery, session: AsyncSession, bot: Bot, state: FSMContext) -> None:
    """Создает чек в CryptoBot и фиксирует выплату (шаг 3 - финальный)."""

    if callback.from_user is None or callback.data is None:
        return

    admin_service = AdminService(session=session)
    if not await admin_service.can_manage_payouts(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return

    admin_user = await UserService(session=session).get_by_telegram_id(callback.from_user.id)
    if admin_user is None:
        await callback.answer("Админ не найден в БД", show_alert=True)
        await state.clear()
        return

    # Получаем данные из FSM
    data = await state.get_data()
    user_id = data.get("payout_user_id")
    total_amount_str = data.get("payout_total_amount", "0.00")
    ledger_page = data.get("payout_ledger_page", 0)
    
    if user_id is None:
        await callback.answer("Ошибка: данные сессии потеряны", show_alert=True)
        await state.clear()
        if callback.message is not None:
            await _edit_payout_ledger_message(callback.message, session, page=int(ledger_page))
        return

    user = await session.get(User, user_id)
    if user is None:
        await callback.answer("Пользователь не найден", show_alert=True)
        await state.clear()
        if callback.message is not None:
            await _edit_payout_ledger_message(callback.message, session, page=int(ledger_page))
        return

    amount = Decimal(total_amount_str)
    if amount <= Decimal("0.00"):
        await callback.answer("Баланс к выплате уже пустой", show_alert=True)
        await state.clear()
        if callback.message is not None:
            await _edit_payout_ledger_message(callback.message, session, page=int(ledger_page))
        return

    try:
        available_usdt = await CryptoBotService().get_available_balance(asset_code="USDT")
    except RuntimeError as exc:
        await callback.answer("Не удалось получить баланс CryptoPay", show_alert=True)
        if callback.message is not None:
            await edit_message_text_safe(
                callback.message,
                (
                    "⚠️ <b>Ошибка проверки баланса CryptoPay</b>\n\n"
                    f"<code>{escape(str(exc))}</code>"
                ),
                reply_markup=payout_final_confirm_keyboard(user_id=user_id, ledger_page=ledger_page),
                parse_mode="HTML",
            )
        return

    if available_usdt < amount:
        await callback.answer("Недостаточно средств в CryptoPay", show_alert=True)
        if callback.message is not None:
            await edit_message_text_safe(
                callback.message,
                (
                    "⚠️ <b>Недостаточно средств в CryptoPay</b>\n\n"
                    f"<b>Нужно:</b> <code>{amount} USDT</code>\n"
                    f"<b>Доступно:</b> <code>{available_usdt} USDT</code>\n\n"
                    "Нажмите «Пополнить через invoice», оплатите счёт и повторите отправку."
                ),
                reply_markup=payout_final_confirm_keyboard(user_id=user_id, ledger_page=ledger_page),
                parse_mode="HTML",
            )
        return

    username = f"@{user.username}" if user.username else f"@{user.telegram_id}"
    comment = f"Payment from @GDPX1 for {username}"
    
    try:
        check = await CryptoBotService().create_usdt_check(amount=amount, comment=comment)
    except RuntimeError as exc:
        error_msg = str(exc)
        # Специальная обработка NOT_ENOUGH_COINS
        if "NOT_ENOUGH_COINS" in error_msg:
            await callback.answer("Недостаточно средств на счёте CryptoBot", show_alert=True)
            if callback.message is not None:
                await edit_message_text_safe(
                    callback.message,
                    f"<b>⚠️ Ошибка CryptoBot:</b>\n{error_msg}\n\n"
                    f"<b>Решение:</b> Пополните баланс CryptoBot и повторите попытку.",
                    reply_markup=payout_final_confirm_keyboard(user_id=user_id, ledger_page=ledger_page),
                    parse_mode="HTML",
                )
            return
        else:
            await callback.answer("Не удалось создать чек CryptoBot", show_alert=True)
            if callback.message is not None:
                await edit_message_text_safe(
                    callback.message,
                    f"Ошибка CryptoBot: {exc}\n\nПопробуйте снова из ведомости.",
                    reply_markup=InlineKeyboardMarkup(
                        inline_keyboard=[
                            [
                                InlineKeyboardButton(
                                    text="💰 К ведомости",
                                    callback_data=f"{CB_PAY_LEDGER_PAGE}:{int(ledger_page)}",
                                )
                            ]
                        ]
                    ),
                )
        await state.clear()
        return

    payout = await BillingService(session=session).mark_user_paid_with_crypto(
        user_id=user_id,
        paid_by_admin_id=admin_user.id,
        crypto_check_id=check.check_id,
        crypto_check_url=check.check_url,
        note="cryptobot_check",
    )
    if payout is None:
        await callback.answer("Выплата уже зафиксирована или баланс нулевой", show_alert=True)
        await state.clear()
        if callback.message is not None:
            await _edit_payout_ledger_message(callback.message, session, page=int(ledger_page))
        return

    await callback.answer("✅ Выплата зафиксирована!")
    await AdminAuditService(session=session).log(
        admin_id=admin_user.id,
        action="mark_paid",
        target_type="user",
        target_id=user_id,
        details=f"amount={payout.amount};check_id={check.check_id}",
    )
    
    # Очищаем состояние после успешной выплаты
    await state.clear()
    
    if callback.message is not None:
        success_text = (
            f"✅ <b>ВЫПЛАТА УСПЕШНО ОТПРАВЛЕНА</b>\n\n"
            f"<b>Сумма:</b> {payout.amount} USDT\n"
            f"<b>Получатель:</b> {username}\n"
            f"<b>Чек:</b> <a href='{check.check_url}'>Открыть чек</a>\n\n"
            f"<i>Возврат в ведомость...</i>"
        )
        await edit_message_text_safe(
            callback.message,
            success_text,
            reply_markup=None,
            parse_mode="HTML",
        )
        # Отправляем обновленную ведомость
        await _edit_payout_ledger_message(callback.message, session, page=ledger_page)
    
    try:
        await bot.send_message(
            user.telegram_id,
            f"✅ Выплата сформирована.\n\nСумма: {payout.amount} USDT\nПолучить чек: {check.check_url}",
        )
    except TelegramAPIError:
        pass

    try:
        await bot.send_message(
            callback.from_user.id,
            f"Выплачено {payout.amount} USDT {username}",
        )
    except TelegramAPIError:
        pass


@router.callback_query(lambda c: c.data is not None and c.data.startswith(f"{CB_PAY_HISTORY_PAGE}:"))
async def on_payout_history_page(callback: CallbackQuery, session: AsyncSession) -> None:
    if callback.from_user is None or callback.data is None:
        return
    if not await AdminService(session=session).can_manage_payouts(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    page = max(int(callback.data.split(":")[2]), 0)
    text, kb = await _payout_history_text_and_markup(session, page=page)
    await callback.answer()
    if callback.message is not None:
        await edit_message_text_safe(callback.message, text, reply_markup=kb)


@router.callback_query(lambda c: c.data is not None and c.data.startswith(f"{CB_PAY_TRASH_PAGE}:"))
async def on_payout_trash_page(callback: CallbackQuery, session: AsyncSession) -> None:
    if callback.from_user is None or callback.data is None:
        return
    if not await AdminService(session=session).can_manage_payouts(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    page = max(int(callback.data.split(":")[2]), 0)
    text, kb = await _payout_trash_text_and_markup(session, page=page)
    await callback.answer()
    if callback.message is not None:
        await edit_message_text_safe(callback.message, text, reply_markup=kb)


@router.callback_query(lambda c: c.data is not None and c.data.startswith(f"{CB_PAY_PENDING_PAGE}:"))
async def on_payout_pending_page(callback: CallbackQuery, session: AsyncSession) -> None:
    if callback.from_user is None or callback.data is None:
        return
    if not await AdminService(session=session).can_manage_payouts(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    page = max(int(callback.data.split(":")[2]), 0)
    text, kb = await _payout_pending_manage_text_and_markup(session, page=page)
    await callback.answer()
    if callback.message is not None:
        await edit_message_text_safe(callback.message, text, reply_markup=kb)


@router.callback_query(lambda c: c.data is not None and c.data.startswith(f"{CB_PAY_PENDING_DELETE}:"))
async def on_payout_pending_delete(callback: CallbackQuery, session: AsyncSession) -> None:
    if callback.from_user is None or callback.data is None:
        return
    if not await AdminService(session=session).can_manage_payouts(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return

    parts = callback.data.split(":")
    if len(parts) < 4:
        await callback.answer("Некорректные данные", show_alert=True)
        return

    try:
        payout_id = int(parts[2])
        page = max(int(parts[3]), 0)
    except ValueError:
        await callback.answer("Некорректные данные", show_alert=True)
        return

    admin_user = await UserService(session=session).get_by_telegram_id(callback.from_user.id)
    if admin_user is None:
        await callback.answer("Админ не найден", show_alert=True)
        return

    deleted = await BillingService(session=session).delete_pending_payout(payout_id=payout_id)
    if deleted is None:
        await callback.answer("PENDING выплата не найдена", show_alert=True)
    else:
        await AdminAuditService(session=session).log(
            admin_id=admin_user.id,
            action="delete_pending_payout",
            target_type="payout",
            target_id=int(deleted["payout_id"]),
            details=f"amount={deleted['amount']};user_id={deleted['user_id']}",
        )
        await callback.answer(f"Удалено: {deleted['amount']} USDT")

    text, kb = await _payout_pending_manage_text_and_markup(session, page=page)
    if callback.message is not None:
        await edit_message_text_safe(callback.message, text, reply_markup=kb)


@router.callback_query(F.data.startswith(f"{CB_ADMIN_REPORT_SUBMISSION}:"))
async def on_submission_report(callback: CallbackQuery, session: AsyncSession) -> None:
    """Показывает детальный отчет по выбранному товару."""

    if callback.from_user is None or callback.data is None:
        return
    if not await AdminService(session=session).is_admin(callback.from_user.id):
        await callback.answer("Недостаточно прав", show_alert=True)
        return

    submission_id = int(callback.data.split(":")[2])
    submission = await session.get(Submission, submission_id)
    if submission is None:
        await callback.answer("Товар не найден", show_alert=True)
        return

    seller = await session.get(User, submission.user_id)
    seller_nickname = f"@{seller.username}" if seller is not None and seller.username else f"@{submission.user_id}"
    category_title = submission.category.title if submission.category is not None else "Без категории"

    actions_stmt = (
        select(ReviewAction).where(ReviewAction.submission_id == submission.id).order_by(ReviewAction.created_at.asc())
    )
    actions = list((await session.execute(actions_stmt)).scalars().all())
    history_lines = [
        f"- {action.created_at}: "
        f"{action.from_status.value if action.from_status else 'none'} -> {action.to_status.value}"
        for action in actions
    ]
    history_text = "\n".join(history_lines) if history_lines else "- без изменений статуса"

    number_line = non_empty_plain((submission.description_text or "").strip(), placeholder="—")
    report_text = (
        f"Отчёт по товару #{submission.id}\n"
        f"Продавец: {seller_nickname}\n"
        f"Категория: {category_title}\n"
        f"📱 `{number_line}` — {category_title}\n"
        f"Текущий статус: {submission_status_emoji_line(submission.status)}\n"
        f"Создано: {submission.created_at}\n"
        f"Взято в работу: {submission.assigned_at}\n"
        f"Проверено: {submission.reviewed_at}\n"
        f"Начислено: {submission.accepted_amount}\n\n"
        "История статусов:\n"
        f"{history_text}"
    )
    if getattr(submission, "is_duplicate", False):
        report_text += "\n⚠️ ВНИМАНИЕ: ЭТОТ НОМЕР УЖЕ БЫЛ В БОТЕ РАНЕЕ!\n"
    await callback.answer("Отчёт сформирован")
    if callback.message is None:
        return

    sent = await callback.message.answer(
        non_empty_plain(report_text) + "\n\n(Сообщение удалится через 20 сек)",
    )
    if sent.chat is not None:
        asyncio.create_task(
            _delete_message_later(
                callback.bot,
                chat_id=sent.chat.id,
                message_id=sent.message_id,
                delay_sec=20,
            )
        )
